import pandas as pd
from openpyxl import Workbook

wb_header_name = ["Secure Deletion Latency", "Data Write Latency", "Journaling Write latency",
                  "Data Read Latency", "Journaling Read latency", "Write Size", "Rewrite Size", "Reread Size",
                  "Read Size", "SD Write Size", "SD Rewrite Size", "SD Reread Size",
                  "dual swap count", "topbuffer (MB)", "scp times"]
wb_item_name = ["CMR", "Native IMR", "BlockSwap", "BlockSwap virtual 1 %", "vg 32MB",
                "vg 64MB", "vg 128MB", "vg 256MB", "vg_reserved 32MB", "vg_reserved 64MB",
                "vg_reserved 128MB", "vg_reserved 256MB", "vg_history 32MB", "vg_history 64MB",
                "vg_history 128MB", "vg_history 256MB"]
traces_name = ["1GB_10GB", "1GB_50GB", "1GB_80GB",
               "10MB_10GB", "10MB_50GB", "10MB_80GB", "TPCC.trace", "TPCE.trace"]
traces_df = []
traces_df_list = []

for trace_name in traces_name:
    traces_df.append(pd.read_csv(
        trace_name + ".csv", delimiter=',', header=None))

for trace_df in traces_df:
    traces_df_list.append(trace_df.values.tolist())

# Create excel
wb = Workbook()


for sheet_name, df_list in zip(traces_name, traces_df_list):
    ws = wb.create_sheet(sheet_name)
    headerRanges = ws['B1':'P1']
    itemRanges = ws['A2':'A17']
    cellRanges = ws['B2':'P17']
    for row in headerRanges:
        for i, c in zip(wb_header_name, row):
            c.value = i
    for i, row in zip(wb_item_name, itemRanges):
        for c in row:
            c.value = i
    for i, row in zip(range(0, 16), cellRanges):
        for j, c in zip(range(0, 15), row):
            c.value = df_list[i][j]

wb.save('IMR_experiment.xlsx')
